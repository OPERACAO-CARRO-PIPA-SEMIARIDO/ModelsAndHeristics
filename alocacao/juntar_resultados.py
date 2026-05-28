import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

PASTA_SAIDAS      = Path(__file__).parent / "saidas_1250_40_365_2"
CAMINHO_XLSX      = PASTA_SAIDAS / "resumo_custos.xlsx"
CAMINHO_INTEGRADO = Path(__file__).parent.parent / "modeloIntegrado" / "resultados00_1250_365" / "historico_controle.csv"

MAPA_INTEGRADO = {
    "abastecimento_00_1250": CAMINHO_INTEGRADO,
}

# Lê sempre do backup CSV (fonte limpa do tester, sem acúmulo de rodadas anteriores)
df_main = pd.read_csv(PASTA_SAIDAS / "backup_temporario.csv", sep=";", decimal=",")

# Lê custos de referência (M1 Diário) por instância
ref_custo = dict(zip(df_main["Nome da Instância"], df_main["Custo M1 Diário"]))

# Monta o bloco do integrado: uma linha por snapshot de tempo
blocos = []
for nome, caminho in MAPA_INTEGRADO.items():
    df_h = pd.read_csv(caminho)
    custo_m1d = ref_custo.get(nome)
    for _, r in df_h.iterrows():
        custo_int = round(float(r["Custo_Roteamento"]), 2)
        gap_mip   = round(float(r["Gap_Percent"]), 4)
        gap_vs    = round(((custo_int - custo_m1d) / custo_m1d) * 100, 2) if custo_m1d else None
        blocos.append({
            "Instancia":              "Modelo Integrado",
            "Hora Limite (h)":        int(r["Hora"]),
            "Tempo Limite (s)":       int(r["Tempo_Segundos"]),
            "Custo Integrado":        custo_int,
            "Gap MIP (%)":            gap_mip,
            "Custo M1 Diário (ref)":  custo_m1d,
            "Gap vs M1 Diário (%)":   gap_vs,
            "Pico Y":                 int(r["Pico_Y"]),
            "Qtd Entregas":           int(r["Qtd_Entregas"]),
        })

df_integrado = pd.DataFrame(blocos)

# Formata números antes de escrever
def _fmt(ws, header_row=1):
    headers = [cell.value for cell in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1):
        for i, cell in enumerate(row):
            if cell.value is None or i >= len(headers):
                continue
            col = str(headers[i])
            if isinstance(cell.value, float):
                if "Custo" in col:
                    cell.number_format = "#,##0.00"
                elif "Gap" in col or "%" in col:
                    cell.number_format = "0.0000"
                elif "Tempo" in col:
                    cell.number_format = "#,##0.0000"
            elif isinstance(cell.value, int):
                if "Total" in col or "Pico" in col or "Qtd" in col or "Hora" in col or "Limite" in col:
                    cell.number_format = "#,##0"
            if "Nome" in col or "Instancia" in col:
                cell.number_format = "@"
                cell.value = str(cell.value)

def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

FILL_HEADER_MAIN = PatternFill("solid", fgColor="4F81BD")
FILL_HEADER_INT  = PatternFill("solid", fgColor="9BBB59")
FONT_HEADER      = Font(bold=True, color="FFFFFF")

with pd.ExcelWriter(CAMINHO_XLSX, engine="openpyxl") as writer:
    # ── Aba principal: dados base ──────────────────────────────────────────
    df_main.to_excel(writer, index=False, sheet_name="Resultados_Custos", startrow=0)
    ws = writer.sheets["Resultados_Custos"]

    # Cabeçalho azul para a seção principal
    for cell in ws[1]:
        cell.fill = FILL_HEADER_MAIN
        cell.font = FONT_HEADER

    _fmt(ws, header_row=1)

    # Linha separadora vazia
    linha_sep = df_main.shape[0] + 3  # +1 header, +1 dados (0-indexed), +1 vazia

    # Cabeçalho verde para a seção do integrado
    cabecalho_int = list(df_integrado.columns)
    for col_idx, titulo in enumerate(cabecalho_int, start=1):
        cell = ws.cell(row=linha_sep, column=col_idx, value=titulo)
        cell.fill = FILL_HEADER_INT
        cell.font = FONT_HEADER

    # Dados do integrado
    for r_idx, row_data in df_integrado.iterrows():
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=linha_sep + 1 + r_idx, column=col_idx, value=val)

    _fmt(ws, header_row=linha_sep)
    _autowidth(ws)

    # ── Aba ModeloIntegrado (progressão completa) ──────────────────────────
    df_integrado.to_excel(writer, index=False, sheet_name="ModeloIntegrado")
    ws2 = writer.sheets["ModeloIntegrado"]
    for cell in ws2[1]:
        cell.fill = FILL_HEADER_INT
        cell.font = FONT_HEADER
    _fmt(ws2, header_row=1)
    _autowidth(ws2)

print(f"Planilha atualizada: {CAMINHO_XLSX}")
